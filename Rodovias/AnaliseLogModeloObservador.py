import re
import json
import time
import requests
import pandas as pd
from pathlib import Path
from datetime import datetime
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


BASE_DIR = Path(__file__).resolve().parent
PASTA_LOGS = BASE_DIR / "LogPrevisoesRodovias"
ARQUIVO_LOG = PASTA_LOGS / "log_previsoes_rodovias.xlsx"
ARQUIVO_RELATORIO = PASTA_LOGS / f"relatorio_metricas_observador_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

REMOVER_DUPLICADOS_MARKET_ID = True
PAUSA_ENTRE_CONSULTAS = 0.10


print("BASE_DIR:", BASE_DIR)
print("PASTA_LOGS:", PASTA_LOGS)
print("ARQUIVO_LOG:", ARQUIVO_LOG)
print("ARQUIVO EXISTE?", ARQUIVO_LOG.exists())


def criar_sessao():
    session = requests.Session()

    retry = Retry(
        total=5,
        connect=5,
        read=5,
        backoff_factor=2,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
    )

    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)

    return session


session = criar_sessao()


def montar_url_mercado(market_id):
    market_id = int(market_id)
    return f"https://app.palpita.io/live/{market_id}-market/rodovia-5-minutos-qu-{market_id}"


def obter_html_mercado(market_id):
    url = montar_url_mercado(market_id)

    headers = {
        "User-Agent": "Mozilla/5.0"
    }

    response = session.get(url, headers=headers, timeout=60)
    response.raise_for_status()

    return response.text


def normalizar_html(html):
    return (
        str(html)
        .replace('\\"', '"')
        .replace("\\/", "/")
        .replace("&quot;", '"')
        .replace("\\u003c", "<")
        .replace("\\u003e", ">")
        .replace("\\u0026", "&")
    )


def extrair_array_json_balanceado(texto, nome_campo):
    texto = normalizar_html(texto)

    pos = texto.find(f'"{nome_campo}"')

    if pos == -1:
        return None

    inicio_array = texto.find("[", pos)

    if inicio_array == -1:
        return None

    nivel = 0
    dentro_string = False
    escape = False

    for i in range(inicio_array, len(texto)):
        char = texto[i]

        if escape:
            escape = False
            continue

        if char == "\\":
            escape = True
            continue

        if char == '"':
            dentro_string = not dentro_string
            continue

        if dentro_string:
            continue

        if char == "[":
            nivel += 1

        elif char == "]":
            nivel -= 1

            if nivel == 0:
                return texto[inicio_array:i + 1]

    return None


def extrair_primeiro_inteiro(texto, padroes):
    texto = normalizar_html(texto)

    for padrao in padroes:
        match = re.search(padrao, texto, re.DOTALL)

        if match:
            try:
                return int(match.group(1))
            except Exception:
                pass

    return None


def extrair_value_final(html):
    padroes = [
        r'"metadata"\s*:\s*\{.*?"valueFinal"\s*:\s*(\d+)',
        r'"valueFinal"\s*:\s*(\d+)',
    ]

    return extrair_primeiro_inteiro(html, padroes)


def extrair_graph_data_points_count(html):
    padroes = [
        r'"graphDataPointsCount"\s*:\s*(\d+)'
    ]

    return extrair_primeiro_inteiro(html, padroes)


def obter_graph_data_mercado(market_id):
    html = obter_html_mercado(market_id)

    texto_array = extrair_array_json_balanceado(
        html,
        "graphData"
    )

    if not texto_array:
        return [], html

    try:
        return json.loads(texto_array), html
    except Exception as e:
        print(f"Erro ao converter graphData para JSON no mercado {market_id}: {e}")
        return [], html


def obter_quantidade_carros_mercado(market_id):
    graph_data, html = obter_graph_data_mercado(market_id)

    value_final = extrair_value_final(html)

    if value_final is not None:
        return value_final, "valueFinal"

    graph_points_count = extrair_graph_data_points_count(html)

    if graph_points_count is not None:
        return graph_points_count, "graphDataPointsCount"

    if graph_data:
        total = 0

        for item in graph_data:
            valor = (
                item.get("value")
                or item.get("v")
                or item.get("y")
                or item.get("count")
                or item.get("carros")
            )

            if valor is None:
                valor = 1

            total += int(valor)

        return total, "graphData"

    return None, "nao_encontrado"


def limpar_texto(valor):
    if pd.isna(valor):
        return ""

    return str(valor).strip()


def obter_direcao_prevista(linha):
    direcao = limpar_texto(linha.get("direcao", "")).upper()

    if direcao in ["MAIS", "ATE"]:
        return direcao

    resultado_previsto = limpar_texto(linha.get("resultado_previsto", "")).lower()

    if resultado_previsto.startswith("mais"):
        return "MAIS"

    if resultado_previsto.startswith("até") or resultado_previsto.startswith("ate"):
        return "ATE"

    return ""


def obter_resultado_real(qtd_carros, meta):
    if qtd_carros is None or meta is None:
        return ""

    if int(qtd_carros) >= int(meta) + 1:
        return "MAIS"

    return "ATE"


def calcular_metricas_df(df, nome_grupo):
    if df.empty:
        return {
            "grupo": nome_grupo,
            "total": 0,
            "acertos": 0,
            "erros": 0,
            "acuracia": 0.0,
            "qtd_mais_previsto": 0,
            "qtd_ate_previsto": 0,
            "qtd_mais_real": 0,
            "qtd_ate_real": 0,
        }

    total = len(df)
    acertos = int(df["acertou"].sum())
    erros = total - acertos
    acuracia = acertos / total if total > 0 else 0

    return {
        "grupo": nome_grupo,
        "total": total,
        "acertos": acertos,
        "erros": erros,
        "acuracia": round(acuracia, 4),
        "qtd_mais_previsto": int((df["direcao_prevista"] == "MAIS").sum()),
        "qtd_ate_previsto": int((df["direcao_prevista"] == "ATE").sum()),
        "qtd_mais_real": int((df["resultado_real"] == "MAIS").sum()),
        "qtd_ate_real": int((df["resultado_real"] == "ATE").sum()),
    }


def gerar_resumo_por_coluna(df, coluna):
    linhas = []

    if df.empty or coluna not in df.columns:
        return pd.DataFrame(linhas)

    for valor in sorted(df[coluna].dropna().unique()):
        if str(valor).strip() == "":
            continue

        grupo = df[df[coluna] == valor]
        linhas.append(calcular_metricas_df(grupo, str(valor)))

    return pd.DataFrame(linhas)


def gerar_resumo_confianca(df):
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["confianca"] = pd.to_numeric(df["confianca"], errors="coerce")

    faixas = [
        (0.00, 0.60, "0.00 até 0.59"),
        (0.60, 0.65, "0.60 até 0.64"),
        (0.65, 0.70, "0.65 até 0.69"),
        (0.70, 0.75, "0.70 até 0.74"),
        (0.75, 1.01, "0.75+"),
    ]

    linhas = []

    for inicio, fim, nome in faixas:
        grupo = df[
            (df["confianca"] >= inicio)
            & (df["confianca"] < fim)
        ]

        linhas.append(calcular_metricas_df(grupo, nome))

    return pd.DataFrame(linhas)


def gerar_resumo_cruzado(df, coluna_1, coluna_2):
    linhas = []

    if df.empty or coluna_1 not in df.columns or coluna_2 not in df.columns:
        return pd.DataFrame(linhas)

    for valor_1 in sorted(df[coluna_1].dropna().unique()):
        for valor_2 in sorted(df[coluna_2].dropna().unique()):
            grupo = df[
                (df[coluna_1] == valor_1)
                & (df[coluna_2] == valor_2)
            ]

            if grupo.empty:
                continue

            metrica = calcular_metricas_df(grupo, f"{valor_1} | {valor_2}")
            metrica[coluna_1] = valor_1
            metrica[coluna_2] = valor_2
            linhas.append(metrica)

    return pd.DataFrame(linhas)


def formatar_excel(caminho_arquivo):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.formatting.rule import CellIsRule

        wb = load_workbook(caminho_arquivo)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                maior = 0
                letra = col[0].column_letter

                for cell in col:
                    valor = "" if cell.value is None else str(cell.value)
                    maior = max(maior, len(valor))

                ws.column_dimensions[letra].width = min(max(maior + 2, 10), 45)

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="center")

            cabecalhos = [cell.value for cell in ws[1]]

            if "acuracia" in cabecalhos:
                col_acuracia = None

                for cell in ws[1]:
                    if cell.value == "acuracia":
                        col_acuracia = cell.column_letter
                        break

                if col_acuracia and ws.max_row >= 2:
                    intervalo = f"{col_acuracia}2:{col_acuracia}{ws.max_row}"

                    for linha in ws[intervalo]:
                        for item in linha:
                            item.number_format = "0.00%"

                    ws.conditional_formatting.add(
                        intervalo,
                        CellIsRule(
                            operator="greaterThanOrEqual",
                            formula=["0.60"],
                            fill=PatternFill("solid", fgColor="C6EFCE"),
                        ),
                    )

                    ws.conditional_formatting.add(
                        intervalo,
                        CellIsRule(
                            operator="lessThan",
                            formula=["0.50"],
                            fill=PatternFill("solid", fgColor="FFC7CE"),
                        ),
                    )

        wb.save(caminho_arquivo)

    except Exception as e:
        print(f"Relatório gerado, mas não foi possível aplicar formatação avançada: {e}")


def salvar_relatorio_excel(df_resultado, df_validos, df_invalidos):
    resumo_geral = pd.DataFrame([
        calcular_metricas_df(df_validos, "GERAL")
    ])

    resumo_direcao = gerar_resumo_por_coluna(df_validos, "direcao_prevista")
    resumo_rodovia = gerar_resumo_por_coluna(df_validos, "rodovia")
    resumo_forca = gerar_resumo_por_coluna(df_validos, "forca_entrada")
    resumo_confianca = gerar_resumo_confianca(df_validos)

    cruzado_rodovia_direcao = gerar_resumo_cruzado(
        df_validos,
        "rodovia",
        "direcao_prevista"
    )

    cruzado_rodovia_forca = gerar_resumo_cruzado(
        df_validos,
        "rodovia",
        "forca_entrada"
    )

    erros = df_validos[df_validos["acertou"] == False].copy() if not df_validos.empty else pd.DataFrame()

    with pd.ExcelWriter(ARQUIVO_RELATORIO, engine="openpyxl") as writer:
        resumo_geral.to_excel(writer, sheet_name="Resumo_Geral", index=False)
        resumo_confianca.to_excel(writer, sheet_name="Por_Confianca", index=False)
        resumo_direcao.to_excel(writer, sheet_name="Por_Direcao", index=False)
        resumo_rodovia.to_excel(writer, sheet_name="Por_Rodovia", index=False)
        resumo_forca.to_excel(writer, sheet_name="Por_Forca", index=False)
        cruzado_rodovia_direcao.to_excel(writer, sheet_name="Rodovia_Direcao", index=False)
        cruzado_rodovia_forca.to_excel(writer, sheet_name="Rodovia_Forca", index=False)
        df_validos.to_excel(writer, sheet_name="Detalhes", index=False)
        erros.to_excel(writer, sheet_name="Erros", index=False)
        df_invalidos.to_excel(writer, sheet_name="Invalidos", index=False)
        df_resultado.to_excel(writer, sheet_name="Resultado_Bruto", index=False)

    formatar_excel(ARQUIVO_RELATORIO)

    print()
    print("=" * 60)
    print("RELATÓRIO EXCEL GERADO")
    print("=" * 60)
    print(ARQUIVO_RELATORIO)


def analisar_log():
    if not ARQUIVO_LOG.exists():
        print(f"Arquivo não encontrado: {ARQUIVO_LOG}")
        return

    df = pd.read_excel(ARQUIVO_LOG)

    if df.empty:
        print("Log vazio.")
        return

    colunas_obrigatorias = ["market_id_aberto", "meta"]

    for coluna in colunas_obrigatorias:
        if coluna not in df.columns:
            print(f"Coluna obrigatória não encontrada no log: {coluna}")
            return

    df = df.copy()

    df = df[df["market_id_aberto"].notna()]
    df = df[df["meta"].notna()]

    if REMOVER_DUPLICADOS_MARKET_ID:
        df = df.drop_duplicates(subset=["market_id_aberto"], keep="last")

    resultados = []
    total_linhas = len(df)

    for posicao, (indice, linha) in enumerate(df.iterrows(), start=1):
        market_id = linha.get("market_id_aberto")
        print(f"\nProcessando {posicao}/{total_linhas} | Mercado {market_id}")

        meta = linha.get("meta")
        direcao_prevista = obter_direcao_prevista(linha)

        if not direcao_prevista:
            resultados.append({
                "market_id": market_id,
                "rodovia": linha.get("rodovia", ""),
                "tag": linha.get("tag", ""),
                "meta": meta,
                "resultado_previsto": linha.get("resultado_previsto", ""),
                "direcao_prevista": "",
                "qtd_carros_real": None,
                "resultado_real": "",
                "acertou": False,
                "confianca": linha.get("confianca", None),
                "forca_entrada": linha.get("forca_entrada", ""),
                "carros_previstos": linha.get("carros_previstos", None),
                "qtd_mercados_usados": linha.get("qtd_mercados_usados", None),
                "origem_resultado": "",
                "status": "SEM_DIRECAO_PREVISTA",
                "erro": "Não foi possível identificar a direção prevista"
            })
            continue

        try:
            qtd_carros, origem = obter_quantidade_carros_mercado(market_id)
            resultado_real = obter_resultado_real(qtd_carros, meta)
            acertou = direcao_prevista == resultado_real if resultado_real else False

            resultados.append({
                "market_id": int(market_id),
                "rodovia": linha.get("rodovia", ""),
                "tag": linha.get("tag", ""),
                "meta": int(meta),
                "resultado_previsto": linha.get("resultado_previsto", ""),
                "direcao_prevista": direcao_prevista,
                "qtd_carros_real": qtd_carros,
                "resultado_real": resultado_real,
                "acertou": acertou,
                "confianca": linha.get("confianca", None),
                "forca_entrada": linha.get("forca_entrada", ""),
                "carros_previstos": linha.get("carros_previstos", None),
                "qtd_mercados_usados": linha.get("qtd_mercados_usados", None),
                "estat_media_simples": linha.get("estat_media_simples", None),
                "estat_media_ponderada_recencia": linha.get("estat_media_ponderada_recencia", None),
                "estat_mediana": linha.get("estat_mediana", None),
                "estat_desvio_padrao": linha.get("estat_desvio_padrao", None),
                "estat_coeficiente_variacao": linha.get("estat_coeficiente_variacao", None),
                "estat_minimo": linha.get("estat_minimo", None),
                "estat_maximo": linha.get("estat_maximo", None),
                "estat_amplitude": linha.get("estat_amplitude", None),
                "estat_tendencia_por_mercado": linha.get("estat_tendencia_por_mercado", None),
                "origem_resultado": origem,
                "status": "OK",
                "erro": ""
            })

            print(
                f"{posicao}/{total_linhas} | "
                f"Mercado {int(market_id)} | "
                f"Meta {int(meta)} | "
                f"Previsto {direcao_prevista} | "
                f"Real {resultado_real} | "
                f"Carros {qtd_carros} | "
                f"{'ACERTOU' if acertou else 'ERROU'}"
            )

            time.sleep(PAUSA_ENTRE_CONSULTAS)

        except Exception as e:
            resultados.append({
                "market_id": market_id,
                "rodovia": linha.get("rodovia", ""),
                "tag": linha.get("tag", ""),
                "meta": meta,
                "resultado_previsto": linha.get("resultado_previsto", ""),
                "direcao_prevista": direcao_prevista,
                "qtd_carros_real": None,
                "resultado_real": "",
                "acertou": False,
                "confianca": linha.get("confianca", None),
                "forca_entrada": linha.get("forca_entrada", ""),
                "carros_previstos": linha.get("carros_previstos", None),
                "qtd_mercados_usados": linha.get("qtd_mercados_usados", None),
                "origem_resultado": "",
                "status": "ERRO_CONSULTA",
                "erro": str(e)
            })

            print(f"Erro ao consultar mercado {market_id}: {e}")

    df_resultado = pd.DataFrame(resultados)

    if df_resultado.empty:
        print("Nenhum resultado processado.")
        return

    df_validos = df_resultado[df_resultado["status"] == "OK"].copy()
    df_invalidos = df_resultado[df_resultado["status"] != "OK"].copy()

    print()
    print("=" * 60)
    print("RESUMO FINAL")
    print("=" * 60)
    print(f"Linhas no log consideradas: {len(df)}")
    print(f"Mercados com resultado consultado: {len(df_validos)}")
    print(f"Mercados com erro/sem resultado: {len(df_invalidos)}")

    if not df_validos.empty:
        resumo = calcular_metricas_df(df_validos, "GERAL")

        print()
        print("GERAL")
        print("-----")
        print(f"Total: {resumo['total']}")
        print(f"Acertos: {resumo['acertos']}")
        print(f"Erros: {resumo['erros']}")
        print(f"Acurácia: {resumo['acuracia'] * 100:.2f}%")

    salvar_relatorio_excel(df_resultado, df_validos, df_invalidos)


if __name__ == "__main__":
    analisar_log()