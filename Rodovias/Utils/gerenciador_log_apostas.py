import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook, load_workbook

# Pega a pasta onde este arquivo .py está salvo
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Cria o caminho para a pasta de logs dentro do projeto
PASTA_LOGS = os.path.join(BASE_DIR, "LogApostasRodovias")

class GerenciadorLogApostas:
    def __init__(self, pasta_base=None, nome_arquivo="log_apostas_rodovias.xlsx"):
        if pasta_base is None:
            pasta_base = PASTA_LOGS

        self.pasta_base = pasta_base
        self.caminho_arquivo = os.path.join(self.pasta_base, nome_arquivo)

        self.abas = {
            "Decisoes": [
                "data_hora",
                "market_id",
                "rodovia",
                "meta_referencia",
                "previsao",
                "confianca",
                "threshold",
                "apostar",
                "motivo"
            ],
            "Ordens": [
                "data_hora",
                "market_id",
                "rodovia",
                "selection_id",
                "order_id",
                "tipo_execucao",
                "odds_tentativa",
                "odd_executada",
                "status_final",
                "valor_total"
            ],
            "Resultados": [
                "data_hora",
                "market_id",
                "rodovia",
                "previsao",
                "resultado_real",
                "acertou",
                "saldo_antes",
                "saldo_depois",
                "lucro_prejuizo"
            ]
        }

        self._garantir_estrutura()

    def _garantir_estrutura(self):
        os.makedirs(self.pasta_base, exist_ok=True)

        if not os.path.exists(self.caminho_arquivo):
            with pd.ExcelWriter(self.caminho_arquivo, engine="openpyxl") as writer:
                for aba, colunas in self.abas.items():
                    df = pd.DataFrame(columns=colunas)
                    df.to_excel(writer, sheet_name=aba, index=False)

    def _safe_load(self):
        if not os.path.exists(self.caminho_arquivo):
            self._garantir_estrutura()

        try:
            return load_workbook(self.caminho_arquivo)
        except Exception:
            # arquivo corrompido → recria corretamente
            print("⚠️ Arquivo de log corrompido. Recriando...")

            if os.path.exists(self.caminho_arquivo):
                os.remove(self.caminho_arquivo)

            self._garantir_estrutura()
            return load_workbook(self.caminho_arquivo)
    
    def _append_linha(self, nome_aba, dados):
        if nome_aba not in self.abas:
            raise ValueError(f"Aba inválida: {nome_aba}")

        dados_final = {}
        for coluna in self.abas[nome_aba]:
            dados_final[coluna] = dados.get(coluna)

        df_novo = pd.DataFrame([dados_final])

        if not os.path.exists(self.caminho_arquivo):
            self._garantir_estrutura()

        workbook = self._safe_load()

        if nome_aba not in workbook.sheetnames:
            with pd.ExcelWriter(self.caminho_arquivo, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df_novo.to_excel(writer, sheet_name=nome_aba, index=False)
            return

        worksheet = workbook[nome_aba]
        startrow = worksheet.max_row

        with pd.ExcelWriter(self.caminho_arquivo, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df_novo.to_excel(writer, sheet_name=nome_aba, index=False, header=False, startrow=startrow)

    def registrar_decisao(
        self,
        market_id,
        rodovia,
        meta_referencia,
        previsao,
        confianca,
        threshold,
        apostar,
        motivo
    ):
        self._append_linha("Decisoes", {
            "data_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "market_id": market_id,
            "rodovia": rodovia,
            "meta_referencia": meta_referencia,
            "previsao": previsao,
            "confianca": confianca,
            "threshold": threshold,
            "apostar": apostar,
            "motivo": motivo
        })

    def registrar_ordem(
        self,
        market_id,
        rodovia,
        selection_id,
        order_id,
        tipo_execucao,
        odds_tentativa,
        odd_executada,
        status_final,
        valor_total
    ):
        self._append_linha("Ordens", {
            "data_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "market_id": market_id,
            "rodovia": rodovia,
            "selection_id": selection_id,
            "order_id": order_id,
            "tipo_execucao": tipo_execucao,
            "odds_tentativa": str(odds_tentativa),
            "odd_executada": odd_executada,
            "status_final": status_final,
            "valor_total": valor_total
        })

    def registrar_resultado(
        self,
        market_id,
        rodovia,
        previsao,
        resultado_real,
        acertou,
        saldo_antes,
        saldo_depois
    ):
        lucro_prejuizo = None
        if saldo_antes is not None and saldo_depois is not None:
            lucro_prejuizo = float(saldo_depois) - float(saldo_antes)

        self._append_linha("Resultados", {
            "data_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "market_id": market_id,
            "rodovia": rodovia,
            "previsao": previsao,
            "resultado_real": resultado_real,
            "acertou": acertou,
            "saldo_antes": saldo_antes,
            "saldo_depois": saldo_depois,
            "lucro_prejuizo": lucro_prejuizo
        })